import os
import openpyxl

def search_files(folder_path, search_word, output_file_path, fileending):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Search Results"

    # Add headers to the table
    sheet.cell(row=1, column=1).value = "File Name"
    sheet.cell(row=1, column=2).value = "Sentence"
    sheet.cell(row=1, column=3).value = "Line Number"

    row = 2  # Start from the second row for data

    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if file.endswith(fileending):
                file_path = os.path.join(root, file)
                try:
                    with open(file_path, "r") as f:
                        lines = f.readlines()
                        for line_number, line in enumerate(lines, start=1):
                            if search_word in line:
                                sheet.cell(row=row, column=1).value = file
                                sheet.cell(row=row, column=2).value = line.strip()
                                sheet.cell(row=row, column=3).value = line_number
                                row += 1
                except IOError:
                    print("Error reading file:", file_path)

    workbook.save(output_file_path)
    print("Excel file created successfully!")

folder_path = r'C:\Users\Tigereye\Desktop\ImportPythonSaveLocations\inputFiles'
search_word = "CreateCloseExcel"
output_file_path = r"C:\Users\Tigereye\Desktop\ImportPythonSaveLocations" + "\\" + search_word + "_outputWordSearch.xlsx"
fileending = ".py"
search_files(folder_path, search_word, output_file_path, fileending)

print('DONE')
